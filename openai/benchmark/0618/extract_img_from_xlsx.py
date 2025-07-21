import re
import requests
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO

input_xlsx = 'openai/benchmark/0618/output.xlsx'
output_xlsx = 'openai/benchmark/0618/output_with_images.xlsx'

# 正则提取 URL
url_pat = re.compile(r'https?://[^\s\)]+')

wb = load_workbook(input_xlsx)
ws = wb.active

# 新增表头
ws.cell(row=1, column=8, value='图片URL')
ws.cell(row=1, column=9, value='图片内容')
col_url, col_img = 8, 9

for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
    ai_resp = row[5].value or ''  # ai_response 在第6列
    m = url_pat.search(ai_resp)
    img_url = m.group(0) if m else ''
    ws.cell(row=i, column=col_url, value=img_url)
    if img_url:
        try:
            resp = requests.get(img_url, timeout=10)
            img = XLImage(BytesIO(resp.content))
            img.width = 100
            img.height = 100
            ws.add_image(img, f'{chr(64+col_img)}{i}')
        except Exception as e:
            ws.cell(row=i, column=col_img, value='图片加载失败')

wb.save(output_xlsx)
print(f'处理完成，结果已保存为 {output_xlsx}') 